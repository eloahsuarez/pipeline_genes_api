from __future__ import annotations

import csv
import json
from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from models import ConservedRegion, PrimerPair, SequenceRecord


def _serializable(value: Any) -> Any:
    if is_dataclass(value):
        return asdict(value)
    if isinstance(value, dict):
        return {key: _serializable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_serializable(item) for item in value]
    return value


def write_text(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8")


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(_serializable(data), ensure_ascii=False, indent=2), encoding="utf-8")


def export_sequences_csv(path: Path, records: Iterable[SequenceRecord]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "selected", "accession", "organism", "length", "definition", "genes",
                "molecule_type", "exons", "sequence",
            ],
        )
        writer.writeheader()
        for record in records:
            writer.writerow(
                {
                    "selected": record.selected,
                    "accession": record.accession,
                    "organism": record.organism,
                    "length": record.length,
                    "definition": record.definition,
                    "genes": "; ".join(record.genes),
                    "molecule_type": record.molecule_type,
                    "exons": "; ".join(
                        f"{exon.number or '?'}:{exon.start}-{exon.end}"
                        for exon in record.exons
                    ),
                    "sequence": record.sequence,
                }
            )


def export_regions_csv(path: Path, regions: Iterable[ConservedRegion]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        fieldnames = list(ConservedRegion.__dataclass_fields__.keys())
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for region in regions:
            writer.writerow(asdict(region))


def export_specificity_csv(path: Path, report: Any) -> None:
    """Exporta o resumo isolado de BLAST+/MFEprimer por par de primers."""

    fieldnames = [
        "rank",
        "parecer",
        "primer_forward",
        "primer_reverse",
        "hits_blast_forward",
        "hits_blast_reverse",
        "produtos_mfeprimer",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for result in report.results:
            writer.writerow(
                {
                    "rank": result.pair_rank,
                    "parecer": result.verdict,
                    "primer_forward": result.forward_sequence,
                    "primer_reverse": result.reverse_sequence,
                    "hits_blast_forward": result.blast_forward_hit_count,
                    "hits_blast_reverse": result.blast_reverse_hit_count,
                    "produtos_mfeprimer": result.mfe_amplicon_count,
                }
            )


def _first_available_attr(value: Any, *names: str, default: Any = "") -> Any:
    """Retorna o primeiro atributo não nulo, preservando zero e ``False``."""

    for name in names:
        candidate = getattr(value, name, None)
        if candidate is not None:
            return candidate
    return default


def _rounded_metric(value: Any) -> Any:
    if value in (None, ""):
        return ""
    try:
        return round(float(value), 3)
    except (TypeError, ValueError):
        return value


def _estimated_hit_differences(hit: Any, *, three_prime: bool) -> Any:
    if hit is None:
        return ""
    if three_prime:
        explicit = _first_available_attr(
            hit,
            "estimated_three_prime_mismatches",
            "estimated_three_prime_differences",
            "estimated_3p_differences",
            "three_prime_differences",
            "differences_3p",
            "three_prime_mismatch_count",
            "mismatches_3p",
            default=None,
        )
        if explicit is not None:
            return explicit
        query_length = getattr(hit, "query_length", None)
        query_end = getattr(hit, "query_end", None)
        if query_length is not None and query_end is not None:
            try:
                return max(int(query_length) - int(query_end), 0)
            except (TypeError, ValueError):
                return ""
        return ""

    explicit = _first_available_attr(
        hit,
        "estimated_mismatches",
        "estimated_total_differences",
        "estimated_differences_total",
        "estimated_differences",
        "total_differences",
        "difference_count",
        "mismatch_count",
        default=None,
    )
    if explicit is not None:
        return explicit
    query_length = getattr(hit, "query_length", None)
    identities = getattr(hit, "identities", None)
    if query_length is not None and identities is not None:
        try:
            return max(int(query_length) - int(identities), 0)
        except (TypeError, ValueError):
            return ""
    return ""


def _product_hits(product: Any) -> tuple[Any, Any, Any, Any]:
    """Obtém sítios espaciais e hits F/R em relatórios novos ou legados."""

    forward = getattr(product, "forward_hit", None)
    reverse = getattr(product, "reverse_hit", None)
    left = getattr(product, "left_hit", None)
    right = getattr(product, "right_hit", None)

    if left is None or right is None:
        candidates = [hit for hit in (forward, reverse) if hit is not None]
        if len(candidates) == 2:
            plus = next(
                (
                    hit
                    for hit in candidates
                    if str(getattr(hit, "subject_strand", "") or "").casefold()
                    == "plus"
                ),
                None,
            )
            minus = next(
                (
                    hit
                    for hit in candidates
                    if str(getattr(hit, "subject_strand", "") or "").casefold()
                    == "minus"
                ),
                None,
            )
            if plus is not None and minus is not None:
                left = left or plus
                right = right or minus
            else:
                ordered = sorted(
                    candidates,
                    key=lambda hit: min(
                        int(getattr(hit, "subject_start", 0) or 0),
                        int(getattr(hit, "subject_end", 0) or 0),
                    ),
                )
                left = left or ordered[0]
                right = right or ordered[1]
        elif len(candidates) == 1:
            left = left or candidates[0]

    # Objetos novos podem fornecer somente left/right. As colunas legadas
    # continuam preenchidas, ainda que em F-F/R-R representem os dois sítios.
    forward = forward or left
    reverse = reverse or right
    return left, right, forward, reverse


def _hit_orientation(hit: Any, *, forward: Any, reverse: Any, fallback: str) -> str:
    orientation = _first_available_attr(
        hit,
        "primer_orientation",
        "orientation",
        default="",
    )
    if orientation:
        return str(orientation).upper()
    if hit is forward and hit is not reverse:
        return "F"
    if hit is reverse and hit is not forward:
        return "R"
    return fallback


def _primer_combination(
    product: Any,
    left: Any,
    right: Any,
    forward: Any,
    reverse: Any,
) -> str:
    explicit = _first_available_attr(
        product,
        "primer_combination",
        "combination",
        default="",
    )
    if explicit:
        return str(explicit).upper()
    left_orientation = _hit_orientation(
        left,
        forward=forward,
        reverse=reverse,
        fallback="F",
    )
    right_orientation = _hit_orientation(
        right,
        forward=forward,
        reverse=reverse,
        fallback="R",
    )
    if left_orientation == right_orientation:
        return f"{left_orientation}-{right_orientation}"
    return "F-R"


def _site_csv_values(prefix: str, hit: Any, orientation: str) -> dict[str, Any]:
    if hit is None:
        return {
            f"{prefix}_orientacao": orientation,
            f"{prefix}_query_cover_pct": "",
            f"{prefix}_identidade_pct": "",
            f"{prefix}_diferencas_estimadas_total": "",
            f"{prefix}_diferencas_estimadas_3p": "",
            f"{prefix}_inicio": "",
            f"{prefix}_fim": "",
            f"{prefix}_fita": "",
        }
    return {
        f"{prefix}_orientacao": orientation,
        f"{prefix}_query_cover_pct": _rounded_metric(
            getattr(hit, "query_coverage_pct", "")
        ),
        f"{prefix}_identidade_pct": _rounded_metric(
            getattr(hit, "identity_pct", "")
        ),
        f"{prefix}_diferencas_estimadas_total": _estimated_hit_differences(
            hit, three_prime=False
        ),
        f"{prefix}_diferencas_estimadas_3p": _estimated_hit_differences(
            hit, three_prime=True
        ),
        f"{prefix}_inicio": getattr(hit, "subject_start", ""),
        f"{prefix}_fim": getattr(hit, "subject_end", ""),
        f"{prefix}_fita": getattr(hit, "subject_strand", ""),
    }


def export_ncbi_specificity_csv(path: Path, report: Any) -> None:
    """Exporta uma linha por produto F-R/F-F/R-R previsto no NCBI."""

    fieldnames = [
        "rank",
        "par",
        "parecer",
        "classificacao",
        "acesso",
        "gene",
        "descricao",
        "organismo",
        "taxid",
        "amplicon_inicio",
        "amplicon_fim",
        "amplicon_bp",
        "combinacao_primers",
        "sitio_esquerdo_orientacao",
        "sitio_esquerdo_query_cover_pct",
        "sitio_esquerdo_identidade_pct",
        "sitio_esquerdo_diferencas_estimadas_total",
        "sitio_esquerdo_diferencas_estimadas_3p",
        "sitio_esquerdo_inicio",
        "sitio_esquerdo_fim",
        "sitio_esquerdo_fita",
        "sitio_direito_orientacao",
        "sitio_direito_query_cover_pct",
        "sitio_direito_identidade_pct",
        "sitio_direito_diferencas_estimadas_total",
        "sitio_direito_diferencas_estimadas_3p",
        "sitio_direito_inicio",
        "sitio_direito_fim",
        "sitio_direito_fita",
        "primer_forward",
        "query_cover_forward_pct",
        "identidade_forward_pct",
        "forward_inicio",
        "forward_fim",
        "forward_fita",
        "primer_reverse",
        "query_cover_reverse_pct",
        "identidade_reverse_pct",
        "reverse_inicio",
        "reverse_fim",
        "reverse_fita",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for result in report.results:
            products = list(result.products)
            if not products:
                writer.writerow(
                    {
                        "rank": result.pair_rank,
                        "par": result.pair_id,
                        "parecer": result.verdict,
                        "primer_forward": result.forward_sequence,
                        "primer_reverse": result.reverse_sequence,
                    }
                )
                continue
            for product in products:
                left, right, forward, reverse = _product_hits(product)
                combination = _primer_combination(
                    product, left, right, forward, reverse
                )
                left_orientation = _hit_orientation(
                    left,
                    forward=forward,
                    reverse=reverse,
                    fallback=combination.split("-", 1)[0],
                )
                right_orientation = _hit_orientation(
                    right,
                    forward=forward,
                    reverse=reverse,
                    fallback=combination.rsplit("-", 1)[-1],
                )
                classification = {
                    "target": "Alvo",
                    "off_target": "Outro gene",
                }.get(product.classification, product.classification)
                row = {
                    "rank": result.pair_rank,
                    "par": result.pair_id,
                    "parecer": result.verdict,
                    "classificacao": classification,
                    "acesso": product.accession,
                    "gene": product.gene,
                    "descricao": product.title,
                    "organismo": product.organism,
                    "taxid": product.taxid if product.taxid is not None else "",
                    "amplicon_inicio": product.start,
                    "amplicon_fim": product.end,
                    "amplicon_bp": product.length,
                    "combinacao_primers": combination,
                    "primer_forward": result.forward_sequence,
                    "query_cover_forward_pct": _rounded_metric(
                        getattr(forward, "query_coverage_pct", "")
                    ),
                    "identidade_forward_pct": _rounded_metric(
                        getattr(forward, "identity_pct", "")
                    ),
                    "forward_inicio": getattr(forward, "subject_start", ""),
                    "forward_fim": getattr(forward, "subject_end", ""),
                    "forward_fita": getattr(forward, "subject_strand", ""),
                    "primer_reverse": result.reverse_sequence,
                    "query_cover_reverse_pct": _rounded_metric(
                        getattr(reverse, "query_coverage_pct", "")
                    ),
                    "identidade_reverse_pct": _rounded_metric(
                        getattr(reverse, "identity_pct", "")
                    ),
                    "reverse_inicio": getattr(reverse, "subject_start", ""),
                    "reverse_fim": getattr(reverse, "subject_end", ""),
                    "reverse_fita": getattr(reverse, "subject_strand", ""),
                }
                row.update(
                    _site_csv_values("sitio_esquerdo", left, left_orientation)
                )
                row.update(
                    _site_csv_values("sitio_direito", right, right_orientation)
                )
                writer.writerow(row)


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for column in ws.columns:
        values = [str(cell.value or "") for cell in column[:50]]
        width = min(max(len(value) for value in values) + 2, 60)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max(width, 10)


def export_primers_xlsx(path: Path, pairs: list[PrimerPair]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pares de primers"
    headers = [
        "Rank", "Score", "Amplicon início", "Amplicon fim", "Amplicon bp",
        "Forward", "F início", "F fim", "F tamanho", "F GC %", "F Tm °C",
        "Reverse", "R início", "R fim", "R tamanho", "R GC %", "R Tm °C",
        "IDT F Tm", "IDT R Tm", "Hairpin F ΔG", "Hairpin R ΔG",
        "Self-dimer F ΔG", "Self-dimer R ΔG", "Heterodimer ΔG",
        "Referência da junção", "F cruza junção", "F junções",
        "R cruza junção", "R junções",
    ]
    ws.append(headers)
    for pair in pairs:
        idt = pair.idt or {}
        f = idt.get("forward", {})
        r = idt.get("reverse", {})
        hetero = idt.get("strongest_hetero_dimer", {})

        def junction_summary(candidate) -> str:
            return "; ".join(
                (
                    f"{match.junction_position} "
                    f"(E{match.left_exon_number or '?'}-E{match.right_exon_number or '?'}; "
                    f"5'={match.primer_5_prime_bases}; 3'={match.primer_3_prime_bases})"
                )
                for match in candidate.junctions
            )

        ws.append(
            [
                pair.rank, round(pair.score, 3), pair.amplicon_start, pair.amplicon_end,
                pair.amplicon_length, pair.forward.sequence, pair.forward.start,
                pair.forward.end, pair.forward.length, round(pair.forward.gc_percent, 2),
                round(pair.forward.tm_c, 2), pair.reverse.sequence, pair.reverse.start,
                pair.reverse.end, pair.reverse.length, round(pair.reverse.gc_percent, 2),
                round(pair.reverse.tm_c, 2),
                f.get("analysis", {}).get("MeltTemp"),
                r.get("analysis", {}).get("MeltTemp"),
                f.get("strongest_hairpin", {}).get("deltaG"),
                r.get("strongest_hairpin", {}).get("deltaG"),
                f.get("strongest_self_dimer", {}).get("DeltaG"),
                r.get("strongest_self_dimer", {}).get("DeltaG"),
                hetero.get("DeltaG"),
                pair.reference_accession,
                pair.forward.spans_exon_junction,
                junction_summary(pair.forward),
                pair.reverse.spans_exon_junction,
                junction_summary(pair.reverse),
            ]
        )
    _style_sheet(ws)

    raw = wb.create_sheet("IDT bruto")
    raw.append(["Rank", "JSON"])
    for pair in pairs:
        raw.append([pair.rank, json.dumps(pair.idt, ensure_ascii=False)])
    _style_sheet(raw)
    wb.save(path)
